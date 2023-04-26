import openpyxl
import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border,Side,Alignment,Font,PatternFill



def query_all(demo,data1,data2,data3,data4,data5,output):
    wb = load_workbook(demo)
    if len(data1)!=0:
        wb = car_num(wb,data1)
    else:
        wb.remove(wb['救護車出勤次數'])
        
    if len(data2)!=0:
        wb = car_avgtime(wb,data2)
    else:
        wb.remove(wb['平均每次救護車出勤時間'])
        
    if len(data3)!=0:
        wb = gohos_peo(wb,data3)
    else:
        wb.remove(wb['急救送醫人次'])
        
    if len(data4)!=0:
        wb = urgent_dispose(wb,data4)
    else:
        wb.remove(wb['緊急救護處置次數'])
    
    if len(data5)!=0:
        wb = level_Triage(wb,data5)
    else:
        wb.remove(wb['院後五級檢傷'])
    
    wb.active = 0
    wb.save(output)



def car_num(wb,data1):
    county={"31":"新北市","01":"臺北市","32":"桃園市","03":"臺中市","05":"臺南市","07":"高雄市","34":"宜蘭縣","33":"新竹縣","35":"苗栗縣","37":"彰化縣","38":"南投縣","39":"雲林縣","40":"嘉義縣","43":"屏東縣",
        "46":"臺東縣","45":"花蓮縣","44":"澎湖縣","11":"基隆市","12":"新竹市","22":"嘉義市","90":"金門縣","91":"連江縣","A1":"基隆港","B1":"臺中港","C1":"高雄港","D1":"花蓮港"}
    timenow = datetime.now().strftime("%Y-%m-%d %H:%M:%S %p")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='right')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font = Font(name='標楷體', size=12)
    totfill = PatternFill(fill_type='solid', fgColor="ffe6cc")
    ws1 = wb['救護車出勤次數']
    wb.active = wb['救護車出勤次數']
    ws1 = wb.active
    start_row1 = ws1.max_row

    #資料過濾
        #刪除縣市代碼之外的key & 僅保留預設欄位的value
    for key in data1.copy().keys():
        if not isinstance(data1[key], list):
            del data1[key]
        elif len(data1[key]) > 2:
            data1[key] = data1[key][:2]

#append縣市和值
    for i, county_code in enumerate(data1):
        row_index = i + 3
        county_name = county[county_code]
        ws1.cell(row=row_index, column=1, value=county_name)
        for j, val in enumerate(data1[county_code]):
            col_index = j + 2
            ws1.cell(row=row_index, column=col_index, value=val)

#新增總計與查詢時間
    end_row1=ws1.max_row
    listOfValues = list(data1.values())
    ws1['A' + str(end_row1+1)].value='總計'
    column = 2+len(listOfValues[0])
    for col in range(2,column):
        char = get_column_letter(col)
        ws1[char + str(end_row1+1)] = f'=SUM({char + str(start_row1)}:{char + str(end_row1)})'
    ws1['A' + str(end_row1+2)].value = '查詢時間：'
    ws1['A' + str(end_row1+2)].alignment = align
    ws1['A' + str(end_row1+2)].font = font
    ws1['B' + str(end_row1+2)].value = timenow
    ws1['B' + str(end_row1+2)].font = font
    
#設定邊框、字型、對齊格式
    for i in range(start_row1,ws1.max_row):
        for j in range(1,ws1.max_column+1):
            ws1.cell(i, j).border = border
            ws1.cell(i, j).alignment = alignment
            ws1.cell(i, j).font = font
#設定填滿格式
    for x in range(len(listOfValues[0])+1):
        ws1[str(chr(65+x)) + str(end_row1+1) ].fill = totfill
    return wb



def car_avgtime(wb,data2):
    county={"31":"新北市","01":"臺北市","32":"桃園市","03":"臺中市","05":"臺南市","07":"高雄市","34":"宜蘭縣","33":"新竹縣","35":"苗栗縣","37":"彰化縣","38":"南投縣","39":"雲林縣","40":"嘉義縣","43":"屏東縣",
        "46":"臺東縣","45":"花蓮縣","44":"澎湖縣","11":"基隆市","12":"新竹市","22":"嘉義市","90":"金門縣","91":"連江縣","A1":"基隆港","B1":"臺中港","C1":"高雄港","D1":"花蓮港"}
    timenow = datetime.now().strftime("%Y-%m-%d %H:%M:%S %p")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='right')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font = Font(name='標楷體', size=12)
    totfill = PatternFill(fill_type='solid', fgColor="ffe6cc")
    ws2 = wb['平均每次救護車出勤時間']
    wb.active = wb['平均每次救護車出勤時間']
    ws2= wb.active
    start_row2=ws2.max_row
    
    #資料過濾
        #刪除縣市代碼之外的key & 僅保留預設欄位的value
    for key in data2.copy().keys():
        if not isinstance(data2[key], list):
            del data2[key]
        elif len(data2[key]) > 5:
            data2[key] = data2[key][:5]
#append資料
    for i, county_code in enumerate(data2):
        row_index = i + 3
        county_name = county[county_code]
        ws2.cell(row=row_index, column=1, value=county_name)
        for j, val in enumerate(data2[county_code]):
            col_index = j + 2
            ws2.cell(row=row_index, column=col_index, value=val)
    
#新增總計與查詢時間
    end_row2=ws2.max_row
    listOfValues = list(data2.values())
    ws2['A' + str(end_row2+1)].value='總平均'
    column = 2+len(listOfValues[0])
    for col in range(2,column):
        char = get_column_letter(col)
        ws2[char + str(end_row2+1)] = f'=AVERAGE({char + str(start_row2+1)}:{char + str(end_row2)})'
    ws2['D' + str(end_row2+2)].value = '查詢時間：'
    ws2['D' + str(end_row2+2)].alignment = align
    ws2['D' + str(end_row2+2)].font = font
    ws2['E' + str(end_row2+2)].value = timenow
    ws2['E' + str(end_row2+2)].font = font

#設定邊框、字型、對齊格式
    for i in range(start_row2,ws2.max_row):
        for j in range(1,ws2.max_column+1):
            ws2.cell(i, j).border = border
            ws2.cell(i, j).alignment = alignment
            ws2.cell(i, j).font = font
#設定填滿格式
    for x in range(len(listOfValues[0])+1):
        ws2[str(chr(65+x)) + str(end_row2+1) ].fill = totfill
        
    return wb



def gohos_peo(wb,data3):
    county={"31":"新北市","01":"臺北市","32":"桃園市","03":"臺中市","05":"臺南市","07":"高雄市","34":"宜蘭縣","33":"新竹縣","35":"苗栗縣","37":"彰化縣","38":"南投縣","39":"雲林縣","40":"嘉義縣","43":"屏東縣",
        "46":"臺東縣","45":"花蓮縣","44":"澎湖縣","11":"基隆市","12":"新竹市","22":"嘉義市","90":"金門縣","91":"連江縣","A1":"基隆港","B1":"臺中港","C1":"高雄港","D1":"花蓮港"}
    timenow = datetime.now().strftime("%Y-%m-%d %H:%M:%S %p")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='right')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font = Font(name='標楷體', size=12)
    totfill = PatternFill(fill_type='solid', fgColor="ffe6cc")
    ws3 = wb['急救送醫人次']
    wb.active = wb['急救送醫人次']
    ws3= wb.active
    start_row3=ws3.max_row
    
    #資料過濾
        #刪除縣市代碼之外的key & 僅保留預設欄位的value
    for key in data3.copy().keys():
        if not isinstance(data3[key], list):
            del data3[key]
        elif len(data3[key]) > 1:
            data3[key] = data3[key][:1]
    
#append資料
    for i, county_code in enumerate(data3):
        row_index = i + 3
        county_name = county[county_code]
        ws3.cell(row=row_index, column=1, value=county_name)
        for j, val in enumerate(data3[county_code]):
            col_index = j + 2
            ws3.cell(row=row_index, column=col_index, value=val)
            
#新增總計與查詢時間
    end_row3=ws3.max_row
    listOfValues = list(data3.values())
    ws3['A' + str(end_row3+1)].value='總計'
    column = 2+len(listOfValues[0])
    for col in range(2,column):
        char = get_column_letter(col)
        ws3[char + str(end_row3+1)] = f'=SUM({char + str(start_row3+1)}:{char + str(end_row3)})'
    ws3['A' + str(end_row3+2)].value = '查詢時間：'
    ws3['A' + str(end_row3+2)].alignment = align
    ws3['A' + str(end_row3+2)].font = font
    ws3['B' + str(end_row3+2)].value = timenow
    ws3['B' + str(end_row3+2)].font = font

#設定邊框、字型、對齊格式
    for i in range(start_row3,ws3.max_row):
        for j in range(1,ws3.max_column+1):
            ws3.cell(i, j).border = border
            ws3.cell(i, j).alignment = alignment
            ws3.cell(i, j).font = font
#設定填滿格式
    for x in range(len(listOfValues[0])+1):
        ws3[str(chr(65+x)) + str(end_row3+1) ].fill = totfill
    return wb



def urgent_dispose(wb,data4):
    county={"31":"新北市","01":"臺北市","32":"桃園市","03":"臺中市","05":"臺南市","07":"高雄市","34":"宜蘭縣","33":"新竹縣","35":"苗栗縣","37":"彰化縣","38":"南投縣","39":"雲林縣","40":"嘉義縣","43":"屏東縣",
        "46":"臺東縣","45":"花蓮縣","44":"澎湖縣","11":"基隆市","12":"新竹市","22":"嘉義市","90":"金門縣","91":"連江縣","A1":"基隆港","B1":"臺中港","C1":"高雄港","D1":"花蓮港"}
    timenow = datetime.now().strftime("%Y-%m-%d %H:%M:%S %p")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='right')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font = Font(name='標楷體', size=12)
    totfill = PatternFill(fill_type='solid', fgColor="ffe6cc")
    ws4 = wb['緊急救護處置次數']
    wb.active = wb['緊急救護處置次數']
    ws4= wb.active
    start_row4=ws4.max_row
    
    #資料過濾
        #刪除縣市代碼之外的key & 僅保留預設欄位的value
    for key in data4.copy().keys():
        if not isinstance(data4[key], list):
            del data4[key]
        elif len(data4[key]) > 5:
            data4[key] = data4[key][:5]
    
#append資料
    for i, county_code in enumerate(data4):
        row_index = i + 3
        county_name = county[county_code]
        ws4.cell(row=row_index, column=1, value=county_name)
        for j, val in enumerate(data4[county_code]):
            col_index = j + 2
            ws4.cell(row=row_index, column=col_index, value=val)
            
#新增總計與查詢時間
    end_row4=ws4.max_row
    listOfValues = list(data4.values())
    ws4['A' + str(end_row4+1)].value='總計'
    column = 2+len(listOfValues[0])
    for col in range(2,column):
        char = get_column_letter(col)
        ws4[char + str(end_row4+1)] = f'=SUM({char + str(start_row4+1)}:{char + str(end_row4)})'
    ws4['D' + str(end_row4+2)].value = '查詢時間：'
    ws4['D' + str(end_row4+2)].alignment = align
    ws4['D' + str(end_row4+2)].font = font
    ws4['E' + str(end_row4+2)].value = timenow
    ws4['E' + str(end_row4+2)].font = font
    
#設定邊框、字型、對齊格式
    for i in range(start_row4,ws4.max_row):
        for j in range(1,ws4.max_column+1):
            ws4.cell(i, j).border = border
            ws4.cell(i, j).alignment = alignment
            ws4.cell(i, j).font = font
#設定填滿格式
    for x in range(len(listOfValues[0])+1):
        ws4[str(chr(65+x)) + str(end_row4+1) ].fill = totfill
    return wb



def level_Triage(wb,data5):
    county={"31":"新北市","01":"臺北市","32":"桃園市","03":"臺中市","05":"臺南市","07":"高雄市","34":"宜蘭縣","33":"新竹縣","35":"苗栗縣","37":"彰化縣","38":"南投縣","39":"雲林縣","40":"嘉義縣","43":"屏東縣",
        "46":"臺東縣","45":"花蓮縣","44":"澎湖縣","11":"基隆市","12":"新竹市","22":"嘉義市","90":"金門縣","91":"連江縣","A1":"基隆港","B1":"臺中港","C1":"高雄港","D1":"花蓮港"}
    timenow = datetime.now().strftime("%Y-%m-%d %H:%M:%S %p")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='right')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font = Font(name='標楷體', size=12)
    totfill = PatternFill(fill_type='solid', fgColor="ffe6cc")
    ws5 = wb['院後五級檢傷']
    wb.active = wb['院後五級檢傷']
    ws5= wb.active
    start_row5=ws5.max_row
    
    #資料過濾
        #刪除縣市代碼之外的key & 僅保留預設欄位的value
    for key in data5.copy().keys():
        if not isinstance(data5[key], list):
            del data5[key]
        elif len(data5[key]) > 5:
            data5[key] = data5[key][:5]
    
#append資料
    for i, county_code in enumerate(data5):
        row_index = i + 3
        county_name = county[county_code]
        ws5.cell(row=row_index, column=1, value=county_name)
        for j, val in enumerate(data5[county_code]):
            col_index = j + 2
            ws5.cell(row=row_index, column=col_index, value=val)
#新增總計與查詢時間
    end_row5=ws5.max_row
    listOfValues = list(data5.values())
    ws5['A' + str(end_row5+1)].value='總計'
    column = 2+len(listOfValues[0])
    for col in range(2,column):
        char = get_column_letter(col)
        ws5[char + str(end_row5+1)] = f'=SUM({char + str(start_row5+1)}:{char + str(end_row5)})'
    ws5['D' + str(end_row5+2)].value = '查詢時間：'
    ws5['D' + str(end_row5+2)].alignment = align
    ws5['D' + str(end_row5+2)].font = font
    ws5['E' + str(end_row5+2)].value = timenow
    ws5['E' + str(end_row5+2)].font = font
            
#設定邊框、字型、對齊格式
    for i in range(start_row5,ws5.max_row):
        for j in range(1,ws5.max_column+1):
            ws5.cell(i, j).border = border
            ws5.cell(i, j).alignment = alignment
            ws5.cell(i, j).font = font
#設定填滿格式
    for x in range(len(listOfValues[0])+1):
        ws5[str(chr(65+x)) + str(end_row5+1) ].fill = totfill
    return wb



data1={
 
}



data2={
    "38": [
        7.8784134248665145,
        5.358100686498855,
        0,
        1.3255148741418763,
        1.1782456140350877,
        2185
    ],
    "39": [
        6.445114205447661,
        4.928704042994992,
        2.4616342982777573,
        1.5221204348357151,
        0.8289788689385611,
        2729,
        1234124
    ],
    "43": [
        6.401124260355029,
        4.985057199211045,
        3.039723865877712,
        1.7427376725838264,
        1.5481577909270217,
        4225,
        124124,
        214124
    ],
    "io":1
}



data3={
    "39": [
        2729,
        124
    ],
    "43": [
        4225,
        547,
        345
    ],
}



data4={
    "38": [
        2185,
        44,
        28,
        26,
        315,
        693,
        23423
    ],
    "39": [
        2729,
        48,
        103,
        92,
        361,
        844
    ],
    "43": [
        4225,
        79,
        61,
        52,
        568,
        1371
    ],
    "io":1
}



data5={
    "38": [
        19,
        26,
        153,
        17,
        5,
        
    ],
    "39": [
        16,
        51,
        236,
        34,
    
    
    ],
    "43": [
        76,
        128,
        521,
        97,
        
    ],"43": [
        76,
        128,
        521,
        97,
        
    ],"44": [
        1,
        28,
        1,
        7,
        
    ],
    "io":1
   
}


query_all('一站通報表.xlsx',data1,data2,data3,data4,data5,'output.xlsx')